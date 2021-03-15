import csv
import itertools
import os
from datetime import datetime

from bokeh.embed import components
from bokeh.layouts import column
from bokeh.models import Title
from bokeh.plotting import figure
from bokeh.resources import CDN
from openpyxl import load_workbook
from bio_diversity import models
from dm_apps import settings


def generate_facility_tank_report(facic_id):
    # figure out the filename
    target_dir = os.path.join(settings.BASE_DIR, 'media', 'temp')
    target_file = "temp_export.xlsx"
    target_file_path = os.path.join(target_dir, target_file)
    target_url = os.path.join(settings.MEDIA_ROOT, 'temp', target_file)

    template_file_path = os.path.join(settings.BASE_DIR, 'bio_diversity', 'static', "report_templates",
                                      "facility_tank_template.xlsx")

    facic = models.FacilityCode.objects.filter(pk=facic_id).get()

    qs = models.Tank.objects.filter(facic_id=facic)

    wb = load_workbook(filename=template_file_path)

    # to order workshees so the first sheet comes before the template sheet, rename the template and then copy the
    # renamed sheet, then rename the copy to template so it exists for other sheets to be created from
    ws = wb['template']
    ws.title = facic.name
    wb.copy_worksheet(ws).title = str("template")
    try:
        ws = wb[facic.name]
    except KeyError:
        print(facic.name, "is not a valid name of a worksheet")

    # start writing data at row 3 in the sheet
    row_count = 3
    for item in qs:

        ws['A' + str(row_count)].value = item.name

        cnt = 0
        year_coll_set = set()
        indv_list, grp_list = item.fish_in_cont()
        if indv_list:
            ws['B' + str(row_count)].value = "Y"
            cnt += len(indv_list)
            year_coll_set |= set([indv.stok_year_coll_str() for indv in indv_list])
        if grp_list:
            for grp in grp_list:
                cnt += grp.fish_in_group()
                year_coll_set |= {grp.__str__()}
        ws['C' + str(row_count)].value = cnt
        ws['D' + str(row_count)].value = str(', '.join(set(year_coll_set)))

        row_count += 1

    wb.save(target_file_path)

    return target_url


def generate_growth_chart(plot_fish):

    if type(plot_fish) == models.Individual:
        len_dets = models.IndividualDet.objects.filter(anidc_id__name="Length").filter(anix_id__indv_id=plot_fish)
        weight_dets = models.IndividualDet.objects.filter(anidc_id__name="Weight").filter(anix_id__indv_id=plot_fish)
    else:
        len_dets = models.IndividualDet.objects.filter(anidc_id__name="Length").filter(anix_id__grp_id=plot_fish)
        weight_dets = models.IndividualDet.objects.filter(anidc_id__name="Weight").filter(anix_id__grp_id=plot_fish)
    
    x_len_data = []
    y_len_data = []
    for len_det in len_dets:
        x_len_data.append(datetime.combine(len_det.detail_date, datetime.min.time()))
        y_len_data.append(len_det.det_val)
        
    x_weight_data = []
    y_weight_data = []
    for weight_det in weight_dets:
        x_weight_data.append(datetime.combine(weight_det.detail_date, datetime.min.time()))
        y_weight_data.append(weight_det.det_val)

    # create a new plot
    title_eng = "Growth Chart for fish"

    p_len = figure(
        tools="pan,box_zoom,wheel_zoom,reset,save",
        x_axis_type='datetime',
        x_axis_label='Date',
        y_axis_label='Length',
        plot_width=600, plot_height=300,
    )
    p_weight = figure(
        tools="pan,box_zoom,wheel_zoom,reset,save",
        x_axis_type='datetime',
        x_axis_label='Date',
        y_axis_label='Weight',
        plot_width=600, plot_height=300,
    )

    p_len.axis.axis_label_text_font_style = 'normal'
    p_weight.axis.axis_label_text_font_style = 'normal'
    p_len.add_layout(Title(text=title_eng, text_font_size="16pt"), 'above')
    p_len.line(x=x_len_data, y=y_len_data, line_width=3)
    p_weight.line(x=x_weight_data, y=y_weight_data, line_width=3)

    #------------------------Data File------------------------------
    target_dir = os.path.join(settings.BASE_DIR, 'media', 'temp')
    target_file = "temp_export.csv"
    target_file_path = os.path.join(target_dir, target_file)
    target_url = os.path.join(settings.MEDIA_ROOT, 'temp', target_file)
    with open(target_file_path, 'w') as data_file:
        writer = csv.writer(data_file)
        writer.writerow(["Date", "Length", "Date", "Weight"])
        writer.writerows(itertools.zip_longest(x_len_data, y_len_data, x_weight_data, y_weight_data))
    scirpt, div = components(column(p_len, p_weight), CDN)
    return scirpt, div, target_url
